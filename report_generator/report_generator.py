# ===============================================================================
# Copyright 2020-2021 Intel Corporation
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ===============================================================================

import argparse
import datetime
import hashlib
import json
from typing import Any, List, Dict
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import openpyxl


def get_property(entry: Dict[str, Any], prop: str):
    keys = prop.split(':')
    value = entry
    for key in keys:
        if key not in value:
            return None
        value = value[key]
    if (not value):
        return "null"
    return value


def xy_to_excel_cell(x: int, y: int) -> str:
    return '{}{}'.format(get_column_letter(x + 1), y + 1)


def get_excel_cell(work_sheet, x: int, y: int):
    return work_sheet[xy_to_excel_cell(x, y)]


def write_cell(
    work_sheet,
    x: int,
    y: int,
    value: str,
    *,
    bold=False,
    number_format='General',
) -> None:
    work_sheet[xy_to_excel_cell(x, y)] = value
    work_sheet[xy_to_excel_cell(x, y)].number_format = number_format
    if bold:
        work_sheet[xy_to_excel_cell(x, y)].font = Font(bold=True)


def is_equal_dict(a: Dict[str, Any], b: Dict[str, Any], props: List[str]) -> bool:
    for prop in props:
        if get_property(a, prop) != get_property(b, prop):
            return False
    return True


def get_metrics(report: Dict[str, Any]) -> List[str]:
    metrics = list()
    was = False
    for i in report:
        if i == "time[s]":
            was = True
            continue
        if was:
            metrics.append(i)
    return metrics


def make_unique(a: List[Any]) -> List[Any]:
    result = list()
    d = dict()
    for i in a:
        if i in d:
            continue
        d[i] = 1
        result.append(i)
    return result


def get_range(
    start_x: int,
    finish_x: int,
    start_y: int,
    finish_y: int,
) -> str:
    return xy_to_excel_cell(start_x, start_y) + ':' + \
           xy_to_excel_cell(finish_x, finish_y)


def can_convert_to_float(string: str) -> bool:
    try:
        float(string)
    except ValueError:
        return False
    return True


def write_aggregation_metric(
    ws,
    write_x: int,
    write_y: int,
    metric_range: str,
    metric_name: str,
) -> None:
    metric_string = '=' + metric_name + '(' + metric_range + ')'
    write_cell(
        ws,
        write_x,
        write_y,
        metric_string,
        number_format='0.00',
    )


def write_header_of_sheet(
    work_sheet,
    algorithm: str,
    header_columns: List[str],
    y_offset: int,
    metrics: List[str],
    agg_offset: int,
    agg_metrics: List[str],
    json_results: List[Dict[str, Any]],
    left_offset: int,
) -> None:
    # write header
    for ind, val in enumerate(header_columns):
        write_cell(work_sheet, ind, y_offset, val.split(':')[-1], bold=True)
    # write aggregation metrics
    if len(json_results) >= 2:
        for ind, val in enumerate(agg_metrics):
            write_cell(
                work_sheet,
                left_offset + len(json_results) - 1,
                agg_offset + ind,
                val,
                bold=True,
            )
    # write names of metrics and jsons
    metric_offset = 0
    json_results_len = len(json_results)
    for metric in metrics:
        write_cell(
            work_sheet,
            left_offset + metric_offset,
            y_offset - 1,
            metric,
            bold=True,
        )
        for json_res in json_results:
            write_cell(
                work_sheet,
                left_offset + metric_offset,
                y_offset,
                json_res["file_name"],
                bold=True,
            )
            metric_offset += 1
        for i in range(json_results_len):
            for j in range(i + 1, json_results_len):
                write_cell(
                    work_sheet,
                    left_offset + metric_offset,
                    y_offset,
                    json_results[i]['file_name'] + ' vs ' + json_results[j]['file_name'],
                    bold=True,
                )
                metric_offset += 1


def get_color_rule(metric: str) -> Any:
    red = 'F85D5E'
    yellow = 'FAF52E'
    green = '58C144'
    if metric in ['geomean', 'time[s]']:
        return ColorScaleRule(
            start_type='num', start_value=0.5, start_color=red,
            mid_type='num', mid_value=1, mid_color=yellow,
            end_type='num', end_value=5, end_color=green)
    if metric == 'average':
        return ColorScaleRule(
            start_type='num', start_value=-3, start_color=red,
            mid_type='num', mid_value=0, mid_color=yellow,
            end_type='num', end_value=3, end_color=green)
    return ColorScaleRule(
        start_type='percentile', start_value=10, start_color=red,
        mid_type='percentile', mid_value=50, mid_color=yellow,
        end_type='percentile', end_value=90, end_color=green)


def get_comparison_method(config: Dict[str, str], metric: str) -> str:
    return config[metric] if metric in config else config['default']


def get_ratio_string(a: str, b: str, comparison_method: str, num_digits=3) -> str:
    splitted_comparison_method = comparison_method.split(' ')
    if splitted_comparison_method[0] == "2":
        a, b = b, a
    return '=ROUND(' + a + splitted_comparison_method[1] + b + f',{num_digits})'


def get_header_parameters(
    json_results: List[Dict[str, Any]],
    full_header_parameters: List[str],
    algorithm: str,
) -> List[str]:
    for json_res in json_results:
        for report in json_res['results']:
            if report['algorithm'] != algorithm:
                continue
            result = list()
            for param in full_header_parameters:
                if get_property(report, param) is not None:
                    result.append(param)
            return result
    raise ValueError(f'There is no {algorithm} in input json(s)')


parser = argparse.ArgumentParser()
parser.add_argument('--result-files', type=str, required=True,
                    help='Benchmark result file names separated by commas')
parser.add_argument('--report-file', type=str,
                    default=f'report_{str(datetime.date.today())}.xlsx')
parser.add_argument('--generation-config', type=str,
                    default='default_report_gen_config.json')
args = parser.parse_args()

# Read input json(s)
json_results: List[Dict[str, Any]] = list()
for file_name in args.result_files.split(','):
    with open(file_name, 'r') as file:
        res = json.load(file)
    res['file_name'] = file_name
    json_results.append(res)

# Read config
with open(args.generation_config, 'r') as file:
    gen_config = json.load(file)

# compute hash for software and hardware configurations
HASH_LIMIT = 8
for i, json_res in enumerate(json_results):
    for ware in ['software', 'hardware']:
        h = hashlib.sha256()
        h.update(bytes(str(json_res[ware]), encoding='utf-8'))
        json_res[f'{ware}_hash'] = h.hexdigest()[:HASH_LIMIT]

# getting metrics for each algorithm
available_algos_and_metrics: Dict[str, List[str]] = dict()
for json_res in json_results:
    for report in json_res['results']:
        metrics: List[str] = get_metrics(report)
        if report['algorithm'] in available_algos_and_metrics:
            available_algos_and_metrics[report['algorithm']] += metrics
        else:
            available_algos_and_metrics[report['algorithm']] = metrics

for ind, val in enumerate(available_algos_and_metrics):
    available_algos_and_metrics[val] = ['time[s]'] + make_unique(available_algos_and_metrics[val])


HEAD_OFFSET = 4
JSON_RESULTS_LEN = len(json_results)

stages: List[str] = [
    'training_preparation',
    'training',
    'computation',
    'prediction_preparation',
    'prediction',
    'alternative_prediction',
    'transformation',
    'search',
    'predict_proba',
]

summary: Dict[str, Dict[str, Dict[str, Dict[str, str]]]] = dict()
wb = openpyxl.Workbook()

for algo in available_algos_and_metrics:
    # algo[:31] because excel warning about length of sheet name no more than 31 symbols
    ws = wb.create_sheet(title=f'{algo[:31]}')
    header_params = get_header_parameters(json_results, gen_config['header'], algo)
    LEFT_OFFSET = len(header_params)
    # writing table header
    for offset, val in enumerate(['file_name', 'software_hash', 'hardware_hash']):
        write_cell(ws, 0, offset, val)
        for i, json_res in enumerate(json_results):
            write_cell(ws, i + 1, offset, json_res[val])

    y_offset = 0
    for stage_key in stages:
        # list of already used results
        used = [
            [False for j in range(len(json_results[i]['results']))]
            for i in range(len(json_results))
        ]
        begin_y_offset = y_offset
        for json_res_ind, json_res in enumerate(json_results):
            for report_ind, report in enumerate(json_res['results']):
                if report['stage'] != stage_key or \
                   report['algorithm'] != algo or \
                   used[json_res_ind][report_ind] is True:
                    continue
                # write parameters
                for offset, config in enumerate(header_params):
                    write_cell(ws, offset, HEAD_OFFSET + 1 + y_offset, get_property(report, config))
                # write all metrics in report
                metric_offset = 0
                for metric in available_algos_and_metrics[algo]:
                    write_cell(
                        ws,
                        LEFT_OFFSET + metric_offset + json_res_ind, HEAD_OFFSET + 1 + y_offset,
                        get_property(report, metric),
                        number_format='0.00',
                    )
                    metric_offset += JSON_RESULTS_LEN * (JSON_RESULTS_LEN + 1) // 2
                used[json_res_ind][report_ind] = True

                # try to find in other configs report with same parameters
                for json_res_comp_ind, json_res_comp in enumerate(json_results[json_res_ind + 1:]):
                    original_index = json_res_ind + 1 + json_res_comp_ind
                    for report_comp_ind, report_comp in enumerate(json_res_comp['results']):
                        if report_comp['stage'] != stage_key or \
                           report_comp['algorithm'] != algo or \
                           used[original_index][report_comp_ind] is True or \
                           not is_equal_dict(report, report_comp, header_params):
                            continue
                        metric_offset = 0
                        for metric in available_algos_and_metrics[algo]:
                            write_cell(
                                ws,
                                LEFT_OFFSET + original_index + metric_offset,
                                HEAD_OFFSET + y_offset + 1,
                                get_property(report_comp, metric),
                                number_format='0.00',
                            )
                            metric_offset += JSON_RESULTS_LEN * (JSON_RESULTS_LEN + 1) // 2
                        used[original_index][report_comp_ind] = True
                y_offset += 1

        if y_offset == begin_y_offset:
            # nothing was written, so do not have to write header & do comparison
            continue
        write_header_of_sheet(
            ws,
            algo,
            header_params,
            HEAD_OFFSET + begin_y_offset,
            available_algos_and_metrics[algo],
            HEAD_OFFSET + y_offset + 1,
            gen_config['aggregation_metrics'],
            json_results,
            LEFT_OFFSET,
        )
        # write aggregation metric & save info for summary
        metric_offset = JSON_RESULTS_LEN
        for metric in available_algos_and_metrics[algo]:
            comparison_offset = 0
            for i in range(JSON_RESULTS_LEN):
                for j in range(i + 1, JSON_RESULTS_LEN):
                    # comprasion
                    for y in range(HEAD_OFFSET + begin_y_offset + 1, HEAD_OFFSET + y_offset + 1):
                        first_offset = LEFT_OFFSET + i + metric_offset - JSON_RESULTS_LEN
                        second_offset = LEFT_OFFSET + j + metric_offset - JSON_RESULTS_LEN
                        first_cell = get_excel_cell(ws, first_offset, y)
                        second_cell = get_excel_cell(ws, second_offset, y)

                        if first_cell.value is None or\
                           second_cell.value is None or \
                           not can_convert_to_float(str(first_cell.value)) or \
                           not can_convert_to_float(str(second_cell.value)):
                            continue
                        write_cell(
                            ws,
                            LEFT_OFFSET + metric_offset + comparison_offset,
                            y,
                            get_ratio_string(
                                xy_to_excel_cell(first_offset, y),
                                xy_to_excel_cell(second_offset, y),
                                get_comparison_method(gen_config['comparison_method'], metric),
                            ),
                            number_format='0.000',
                        )
                    # fill comparison range by color rule
                    ws.conditional_formatting.add(
                        get_range(
                            LEFT_OFFSET + metric_offset + comparison_offset,
                            LEFT_OFFSET + metric_offset + comparison_offset,
                            HEAD_OFFSET + 1 + begin_y_offset,
                            HEAD_OFFSET + y_offset,
                        ),
                        get_color_rule(metric),
                    )
                    # write aggregation metric
                    for agg_offset, agg_metric in enumerate(gen_config['aggregation_metrics']):
                        write_aggregation_metric(
                            ws,
                            LEFT_OFFSET + metric_offset + comparison_offset,
                            HEAD_OFFSET + 1 + y_offset + agg_offset,
                            get_range(
                                LEFT_OFFSET + metric_offset + comparison_offset,
                                LEFT_OFFSET + metric_offset + comparison_offset,
                                HEAD_OFFSET + 1 + begin_y_offset,
                                HEAD_OFFSET + y_offset,
                            ),
                            agg_metric,
                        )

                        column_name = \
                            json_results[i]['file_name'] + \
                            ' vs ' + \
                            json_results[j]['file_name'] + \
                            ' (' + stage_key + ')'

                        cell_name_to_summary = \
                            '=' + algo[:31] + '!' + \
                            xy_to_excel_cell(LEFT_OFFSET + metric_offset + comparison_offset,
                                             HEAD_OFFSET + 1 + y_offset + agg_offset)
                        if agg_metric not in summary:
                            summary[agg_metric] = dict()
                        if column_name not in summary[agg_metric]:
                            summary[agg_metric][column_name] = dict()
                        if algo not in summary[agg_metric][column_name]:
                            summary[agg_metric][column_name][algo] = dict()
                        summary[agg_metric][column_name][algo].update(
                            {f'{metric}': cell_name_to_summary})
                    comparison_offset += 1
            metric_offset += JSON_RESULTS_LEN * (JSON_RESULTS_LEN + 1) // 2
        # for comfortable view
        y_offset += len(gen_config['aggregation_metrics']) + 3

# write summary for each aggregation metric
for agg_metric in gen_config['aggregation_metrics']:
    if JSON_RESULTS_LEN == 1:
        continue
    y_offset = 0
    # write summary
    ws = wb.create_sheet('Summary' + f' ({agg_metric})', 0)
    for name_ind, name in enumerate(summary[agg_metric]):
        # write table name
        write_cell(ws, 0, y_offset, name, bold=True)
        # getting unique list of metrics on current comparison
        metrics_in_current_summary = list()
        for algo in summary[agg_metric][name]:
            for metric in summary[agg_metric][name][algo]:
                metrics_in_current_summary.append(metric)
        metrics_in_current_summary = make_unique(metrics_in_current_summary)

        # fill table
        for metric_ind, metric in enumerate(metrics_in_current_summary):
            # write metric name
            write_cell(ws, metric_ind + 1, y_offset + 1, metric)
            for algo_ind, algo in enumerate(summary[agg_metric][name]):
                if metric not in summary[agg_metric][name][algo]:
                    continue
                # write algorithm name
                write_cell(
                    ws,
                    0,
                    y_offset + algo_ind + 2,
                    algo
                )
                # write geomean
                write_cell(
                    ws,
                    metric_ind + 1,
                    y_offset + algo_ind + 2,
                    summary[agg_metric][name][algo][metric],
                    number_format='0.00',
                )

        # color some range by color rule
        ws.conditional_formatting.add(
            get_range(
                1,
                len(metrics_in_current_summary),
                y_offset + 2,
                y_offset + len(summary[agg_metric][name]) + 1,
            ),
            get_color_rule(agg_metric),
        )
        y_offset += len(summary[agg_metric][name]) + 3

# write hardware & software configs
for i, json_res in enumerate(json_results):
    ws = wb.create_sheet(title=f"SW config n{i}_{json_res['software_hash']}")
    ws[xy_to_excel_cell(0, 0)] = \
        f"Software configuration {i} (hash: {json_res['software_hash']})"
    sw_conf = json.dumps(json_res['software'], indent=4).split('\n')
    for j, val in enumerate(sw_conf):
        ws[xy_to_excel_cell(0, 1 + j)] = val

    ws = wb.create_sheet(title=f"HW config n{i}_{json_res['hardware_hash']}")
    ws[xy_to_excel_cell(0, 0)] = \
        f"Hardware configuration {i} (hash: {json_res['hardware_hash']})"
    hw_conf = json.dumps(json_res['hardware'], indent=4).split('\n')
    for j, val in enumerate(hw_conf):
        ws[xy_to_excel_cell(0, 1 + j)] = val

wb.remove(wb['Sheet'])
wb.save(args.report_file)
