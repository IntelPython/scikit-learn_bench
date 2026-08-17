# ===============================================================================
# Copyright 2024 Intel Corporation
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ===============================================================================

import argparse
import json
from typing import Dict, List

import numpy as np
import openpyxl as xl
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy.stats import gmean

try:
    import matplotlib.pyplot as plt
    from matplotlib.ticker import NullLocator, FixedFormatter, FixedLocator
    matplotlib_available = True
except ImportError:
    matplotlib_available = False

from ..utils.common import custom_format, flatten_list
from ..utils.logger import logger
from ..utils.measurement import enrich_metrics
from .compatibility import transform_results_to_compatible

METRICS = {
    "lower is better": [
        "1st run time[ms]",
        "median time[ms]",
        "time[ms]",
        "cost[microdollar]",
        "iterations",
        # classification
        "logloss",
        # regression
        "RMSE",
        # clustering
        "inertia",
        "Davies-Bouldin score",
        # manifold - TSNE
        "Kullback-Leibler divergence",
    ],
    "higher is better": [
        "throughput[samples/ms]",
        # classification
        "accuracy",
        "balanced accuracy",
        "ROC AUC",
        # regression
        "R2",
        # clustering
        "homogeneity",
        "completeness",
        # search
        "recall@10",
    ],
    "indifferent": [
        # SVM
        "support vectors",
        # PCA
        "average log-likelihood",
        "1st component variance ratio",
        # DBSCAN
        # NB: 'n_clusters' is parameter of KMeans while
        # 'clusters' is number of computer clusters by DBSCAN
        "clusters",
    ],
    "incomparable": [
        "1st-mean run ratio",
        "time CV",
        "cpu load[%]",
    ],
}
MEMORY_TYPES = ["RAM", "VRAM"]
for memory_type in MEMORY_TYPES:
    METRICS["incomparable"].append(f"peak {memory_type} usage[MB]")
    METRICS["incomparable"].append(f"{memory_type} usage-iteration correlation")
METRIC_NAMES = flatten_list([list(METRICS[key]) for key in METRICS])
PERF_METRICS = ["time[ms]", "throughput[samples/ms]", "cost[microdollar]"]

COLUMNS_ORDER = [
    # algorithm
    "stage",
    "task",
    "library",
    "estimator",
    "method",
    "function",
    "device",
    "environment_name",
    # data
    "dataset",
    "samples",
    "features",
    "format",
    "dtype",
    "order",
    "n_classes",
    "n_clusters",
    "batch_size",
]

RED_COLOR, YELLOW_COLOR, GREEN_COLOR, WHITE_COLOR = "F85D5E", "FAF52E", "58C144", "FFFFFF"
COLUMN_COLOR_RULES = {
    "time CV": ColorScaleRule(
        start_type="num",
        start_value=0.0,
        start_color=GREEN_COLOR,
        mid_type="num",
        mid_value=0.1,
        mid_color=YELLOW_COLOR,
        end_type="num",
        end_value=0.5,
        end_color=RED_COLOR,
    )
}

DIFFBY_COLUMNS = ["environment_name", "library", "format", "device"]


def geomean_wrapper(a):
    return gmean(a, nan_policy="omit")


def reorder_columns(input_columns: List, columns_order: List = COLUMNS_ORDER) -> List:
    output_columns = list()
    # 1st step: select existing columns from known ordered columns
    for ordered_column in columns_order:
        if ordered_column in input_columns:
            output_columns.append(ordered_column)
            input_columns.remove(ordered_column)
    # 2nd step: add left input columns
    output_columns += input_columns
    return output_columns


def filter_nan_columns(input_df: pd.DataFrame):
    output_df = input_df.copy()
    non_nan_columns = output_df.columns[output_df.isna().mean(axis=0) < 1]
    output_df = output_df[non_nan_columns]
    return output_df


def split_df_by_columns(
    input_df: pd.DataFrame, columns: List, remove_column: bool = True
) -> Dict[str, pd.DataFrame]:
    split_columns = list(set(columns) & set(input_df.columns))
    split_columns = reorder_columns(split_columns, columns)
    value_counts = input_df.value_counts(split_columns, dropna=False, sort=False)
    output_dfs = {}
    for unique_values in value_counts.index:
        index_mask = [
            input_df[column] == unique_value
            for column, unique_value in zip(value_counts.index.names, unique_values)
            if not pd.isna(unique_value)
        ]
        index_mask = pd.DataFrame(index_mask).all(axis=0)
        subset_name = str(unique_values)[1:-1]
        subset_name = subset_name.replace(", ", "|").replace(",", "").replace("'", "")
        subset_name = subset_name.replace("nan|", "").replace("|nan", "")
        output_dfs[subset_name] = filter_nan_columns(input_df.loc[index_mask])
        if remove_column:
            output_dfs[subset_name] = output_dfs[subset_name].drop(
                columns=set(split_columns) & set(output_dfs[subset_name].columns)
            )
        output_dfs[subset_name] = output_dfs[subset_name][
            reorder_columns(list(output_dfs[subset_name].columns))
        ]
    return output_dfs


def compare_df(input_df, diff_columns, diffs_selection, compared_columns=METRIC_NAMES):
    def select_comparison(i, j, diffs_selection):
        if diffs_selection == "upper_triangle":
            return j > i
        elif diffs_selection == "lower_triangle":
            return i > j
        return i != j

    index_columns = list(
        (set(input_df.columns) - set(diff_columns)) - set(compared_columns)
    )
    df = input_df.set_index(index_columns)
    unique_indices = df.index.unique()
    splitted_dfs = split_df_by_columns(input_df, diff_columns)
    for key, df in splitted_dfs.items():
        for index_column in index_columns:
            if index_column not in df.columns:
                df[index_column] = np.nan
    splitted_dfs = {key: df.set_index(index_columns) for key, df in splitted_dfs.items()}

    # drop results with duplicated indices (keep first entry only)
    for key, splitted_df in splitted_dfs.items():
        splitted_dfs[key] = splitted_df[~splitted_df.index.duplicated(keep="first")]

    df = pd.DataFrame(index=unique_indices)
    # original values
    for key, splitted_df in splitted_dfs.items():
        if len(set(splitted_df.columns) - set(compared_columns)) > 0:
            raise ValueError
        for column in splitted_df.columns:
            df[f"{key}\n{column}"] = splitted_df[column]
    # compared values
    for i, (key_ith, df_ith) in enumerate(splitted_dfs.items()):
        for j, (key_jth, df_jth) in enumerate(splitted_dfs.items()):
            if select_comparison(i, j, diffs_selection):
                comparison_name = f"{key_jth} vs {key_ith}"
                for column in df_ith.columns:
                    if column not in df_jth.columns:
                        continue
                    if column in METRICS["higher is better"]:
                        df[f"{comparison_name}\n{column} relative improvement"] = (
                            df_jth[column] / df_ith[column]
                        )
                    elif column in METRICS["lower is better"]:
                        df[f"{comparison_name}\n{column} relative improvement"] = (
                            df_ith[column] / df_jth[column]
                        )
                    elif column in METRICS["indifferent"]:
                        df[f"{comparison_name}\n{column} is equal"] = df_ith[column].eq(
                            df_jth[column]
                        )
    df = df.reset_index()
    # move to multi-index
    df = df[reorder_columns(list(df.columns))]
    df.columns = [
        column if "\n" in column else f"parameter\n{column}" for column in df.columns
    ]
    df.columns = pd.MultiIndex.from_tuples(
        [tuple(column.split("\n")) for column in df.columns]
    )
    return df


def write_df_to_sheet(df, sheet, index=True, header=True):
    for row in dataframe_to_rows(df, index=index, header=header):
        if any(map(lambda x: x is not None, row)):
            sheet.append(row)


def merge_result_files(filenames):
    results = dict()
    for result_name in filenames:
        with open(result_name, "r") as fp:
            result = json.load(fp)
        for key, value in result.items():
            if key in results:
                if isinstance(value, list):
                    results[key] += value
                elif isinstance(value, dict):
                    results[key].update(value)
            else:
                results[key] = value
    return results


def get_result_tables_as_df(
    results,
    diffby_columns=DIFFBY_COLUMNS,
    splitby_columns=["estimator", "method", "function"],
    compatibility_mode=False,
    include_performance_stability_metrics=False,
):
    bench_cases = pd.DataFrame(
        [
            enrich_metrics(bench_case, include_performance_stability_metrics)
            for bench_case in results["bench_cases"]
        ]
    )

    if compatibility_mode:
        bench_cases = transform_results_to_compatible(bench_cases)

    for column in diffby_columns.copy():
        if bench_cases[column].nunique() == 1:
            bench_cases.drop(columns=[column], inplace=True)
            diffby_columns.remove(column)

    return split_df_by_columns(bench_cases, splitby_columns)


def get_summary_from_df(df: pd.DataFrame, df_name: str) -> pd.DataFrame:
    metric_columns = list()
    for column in list(df.columns):
        for metric_name in METRIC_NAMES:
            # only relative improvements are included in summary currently
            if len(column) > 1 and column[1] == f"{metric_name} relative improvement":
                metric_columns.append(column)
    summary = df[metric_columns].aggregate(geomean_wrapper, axis=0).to_frame().T
    summary.index = pd.Index([df_name])
    return summary


def get_color_rule_for_comparison(scale):
    start_value, mid_value, end_value = scale
    return ColorScaleRule(
        start_type="num",
        start_value=start_value,
        start_color=RED_COLOR,
        mid_type="num",
        mid_value=mid_value,
        mid_color=WHITE_COLOR,
        end_type="num",
        end_value=end_value,
        end_color=GREEN_COLOR,
    )


def apply_rules_for_sheet(sheet, perf_color_scale, quality_color_scale):
    for column in sheet.iter_cols():
        column_idx = get_column_letter(column[0].column)
        cell_range = f"${column_idx}1:${column_idx}{len(column)}"
        is_rel_impr = any(
            [
                isinstance(cell.value, str) and "relative improvement" in cell.value
                for cell in column
            ]
        )
        is_perf = any(
            [
                isinstance(cell.value, str)
                and (any(map(lambda x: x in cell.value, PERF_METRICS)))
                for cell in column
            ]
        )
        if is_rel_impr:
            sheet.conditional_formatting.add(
                cell_range,
                get_color_rule_for_comparison(
                    perf_color_scale if is_perf else quality_color_scale
                ),
            )
        else:
            column_name = {cell.value for cell in column} & set(COLUMN_COLOR_RULES.keys())
            if len(column_name) == 1:
                column_name = column_name.pop()
                sheet.conditional_formatting.add(
                    cell_range, COLUMN_COLOR_RULES[column_name]
                )


def prepare_all_cases_df(all_cases_df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepare All cases dataframe with specific column ordering:
    1. Algorithm name (df_name from multi-index)
    2. Parameters and other columns
    3. time[ms] related columns
    4. Exclude metric columns (except time[ms])
    """
    df = all_cases_df.copy()
    
    # Flatten multi-index columns for easier processing
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = ["|".join(col).strip() for col in df.columns.values]
    
    # Identify column groups
    algorithm_cols = [col for col in df.columns if col.startswith("parameter|")]
    time_cols = [col for col in df.columns if "time[ms]" in col and "parameter" not in col]
    
    # Get all metric columns to exclude (except time[ms])
    metric_cols_to_exclude = [
        col for col in df.columns
        if any(metric in col for metric in METRIC_NAMES)
        and "time[ms]" not in col
        and "parameter" not in col
    ]
    
    # Get remaining columns (parameters) - exclude metrics
    remaining_cols = [
        col for col in df.columns
        if col not in algorithm_cols
        and col not in time_cols
        and col not in metric_cols_to_exclude
    ]
    
    ordered_cols = remaining_cols + time_cols + algorithm_cols 
    
    # Filter to only include columns that exist
    ordered_cols = [col for col in ordered_cols if col in df.columns]
    # Select only the ordered columns
    df = df[ordered_cols]
    
    # Convert back to multi-index columns if there were multi-index columns
    if "|" in ordered_cols[0] if ordered_cols else False:
        df.columns = pd.MultiIndex.from_tuples(
            [tuple(col.split("|")) for col in df.columns]
        )
    
    return df


def write_all_cases_2_sheet(dfs, wb):
    """
    Write 'All cases' sheet: one block per algorithm group with columns
      Algorithm | sklearn time[ms] | sklearnex time[ms] | Speedup | <stability> | <parameters>.
    KNN groups are split by method (brute/kd_tree); rows are sorted by dtype then dataset.
    Each block ends with per-dtype and total GEOMEAN formula rows plus a speedup color scale.
    Returns list of (group_name, total_geomean_row, {dtype: geomean_row}) for the summary sheet.
    """
    KNN = ("KNeighborsClassifier", "KNeighborsRegressor")
    STABILITY = ["1st run time[ms]", "median time[ms]", "time CV"]
    ws = wb.create_sheet(title="All cases", index=1)
    row, geomean_rows = 1, []

    def geomean_row(label, r0, r1):
        nonlocal row
        ws.append([label] + [f"=GEOMEAN({get_column_letter(c)}{r0}:{get_column_letter(c)}{r1})"
                             for c in (2, 3, 4)])
        row += 1
        return row - 1

    for df_name, df in dfs.items():
        if not isinstance(df.columns, pd.MultiIndex):
            continue
        time_cols = {p: (p, "time[ms]") for p in ("sklearn", "sklearnex")
                     if (p, "time[ms]") in df.columns}
        speedup = next((c for c in df.columns if "time[ms] relative improvement" in c[1]), None)
        if len(time_cols) < 2 or speedup is None:
            continue

        is_knn = df_name.split("|")[0] in KNN
        params = [c for c in df.columns
                  if c[0] == "parameter" and not (is_knn and c[1] == "algorithm")]
        stability = [(p, m) for m in STABILITY for p in ("sklearn", "sklearnex")
                     if (p, m) in df.columns]
        value_cols = [time_cols["sklearn"], time_cols["sklearnex"], speedup] + stability + params

        algo_col = ("parameter", "algorithm")
        if is_knn and algo_col in df.columns:
            groups = {f"{df_name} ({v})": df[df[algo_col] == v]
                      for v in df[algo_col].dropna().unique()}
        else:
            groups = {df_name: df}

        for name, gdf in groups.items():
            sort_cols = [c for c in (("parameter", "dtype"), ("parameter", "dataset"))
                         if c in gdf.columns]
            if sort_cols:
                gdf = gdf.sort_values(sort_cols, kind="mergesort")

            header = (["Algorithm", "sklearn time[ms]", "sklearnex time[ms]", "Speedup"]
                      + [f"{p} {m}" for p, m in stability] + [c[1] for c in params])
            ws.append(header)
            for cell in ws[row]:
                cell.alignment = Alignment(wrap_text=True)
            row += 1

            start = row
            for _, r in gdf.iterrows():
                ws.append([name] + [r.get(c) for c in value_cols])
                row += 1
            end = row - 1

            if end >= start:
                # per-dtype GEOMEAN rows (data is dtype-sorted, so each dtype is contiguous)
                dtype_rows = {}
                dtype_col = ("parameter", "dtype")
                if dtype_col in gdf.columns:
                    dtypes = gdf[dtype_col].astype(str).tolist()
                    i = 0
                    while i < len(dtypes):
                        j = i
                        while j + 1 < len(dtypes) and dtypes[j + 1] == dtypes[i]:
                            j += 1
                        dtype_rows[dtypes[i]] = geomean_row(f"GEOMEAN {dtypes[i]}", start + i, start + j)
                        i = j + 1

                total_row = geomean_row("GEOMEAN total", start, end)
                geomean_rows.append((name, total_row, dtype_rows))

                vals = gdf[speedup].dropna()
                if len(vals):
                    lo, hi = float(vals.min()), float(vals.max())
                    ws.conditional_formatting.add(
                        f"$D${start}:$D${total_row}",
                        ColorScaleRule(
                            start_type="num", start_value=lo, start_color=RED_COLOR,
                            mid_type="num", mid_value=(lo + hi) / 2, mid_color=YELLOW_COLOR,
                            end_type="num", end_value=hi, end_color=GREEN_COLOR))

            ws.append([])
            row += 1

    return geomean_rows


def write_summary_2_sheet(geomean_rows, wb):
    """
    Write 'Summary (for plots)' sheet with columns:
      Algorithm | geomean sklearn | geomean sklearnex | geomean speedup
    Values are cell references to the GEOMEAN rows in 'All cases'.
    Includes an overall GEOMEAN of speedups and conditional formatting.
    """
    src_sheet_name = "'All cases'"
    ws = wb.create_sheet(title="Summary (for plots)", index=0)

    # Separate into training and inference groups
    training_rows = [(n, r, d) for n, r, d in geomean_rows
                     if "|fit" in n or n == "train_test_split"]
    inference_rows = [(n, r, d) for n, r, d in geomean_rows
                      if (n, r, d) not in training_rows]

    # Columns: A=algo, B/C/D=total, E/F/G=fp32, H/I/J=fp64
    # Each triplet: sklearn time, sklearnex time, speedup
    speedup_cols = ["D", "G", "J"]
    current_row = 1

    HEADER = ["Algorithm",
              "sklearn time[ms]", "sklearnex time[ms]", "speedup",
              "sklearn fp32", "sklearnex fp32", "speedup fp32",
              "sklearn fp64", "sklearnex fp64", "speedup fp64"]

    def refs_for_row(row_num):
        if row_num is None:
            return [None, None, None]
        return [f"={src_sheet_name}!{c}{row_num}" for c in "BCD"]

    def write_section(title, rows):
        nonlocal current_row
        ws.append([title] + HEADER[1:])
        current_row += 1
        start = current_row
        for name, total_row, dtype_rows in rows:
            ws.append([name]
                      + refs_for_row(total_row)
                      + refs_for_row(dtype_rows.get("float32"))
                      + refs_for_row(dtype_rows.get("float64")))
            current_row += 1
        end = current_row - 1
        for col in speedup_cols:
            ws.conditional_formatting.add(
                f"${col}${start}:${col}${end}",
                ColorScaleRule(
                    start_type="min", start_color=RED_COLOR,
                    mid_type="percentile", mid_value=50, mid_color=YELLOW_COLOR,
                    end_type="max", end_color=GREEN_COLOR,
                ))
        return start, end

    t_start, t_end = write_section("Training", training_rows)
    i_start, i_end = write_section("Inference", inference_rows)

    # Summary GEOMEANs (uncolored)
    ws.append([])
    current_row += 1
    for label, ranges in [
        ("Training GEOMEAN", [(t_start, t_end)]),
        ("Inference GEOMEAN", [(i_start, i_end)]),
        ("Total GEOMEAN", [(t_start, t_end), (i_start, i_end)]),
    ]:
        def gm(col):
            return f"=GEOMEAN({','.join(f'{col}{s}:{col}{e}' for s, e in ranges)})"
        ws.append([label, None, None, gm("D"), None, None, gm("G"), None, None, gm("J")])
        current_row += 1


def write_environment_info(results, workbook):
    env_infos = results["environment"]
    for env_name, env_info in env_infos.items():
        for info_type, info_subclass in env_info.items():
            new_ws = workbook.create_sheet(title=f"{info_type}|{env_name}"[:31])
            for sub_key, sub_info in info_subclass.items():
                if isinstance(sub_info, dict):
                    if all(
                        map(
                            lambda x: not (isinstance(x, list) or isinstance(x, dict)),
                            sub_info.values(),
                        )
                    ):
                        info_df = pd.Series(sub_info).to_frame()
                    else:
                        info_df = pd.DataFrame(sub_info).T
                elif isinstance(sub_info, list):
                    info_df = pd.DataFrame(sub_info)
                else:
                    continue
                write_df_to_sheet(info_df, new_ws)
                new_ws.append([None])


def draw_summary_plots(all_cases_df: pd.DataFrame, output_file: str = None):
    """
    Draw plots from all_cases dataframe with algorithm comparison data.
    Separates into fit (training) and predict (inference) plots.
    Calculates geometric means of improvement for each group.
    Separates KNN algorithms by parameter|algorithm value (brute force vs kd_tree).
    """
    if not matplotlib_available:
        logger.warning("matplotlib is not available, skipping plot generation")
        return
    
    try:
        # Flatten column names if multi-index
        if isinstance(all_cases_df.columns, pd.MultiIndex):
            all_cases_df.columns = ["|".join(col).strip() for col in all_cases_df.columns.values]

        algo_name_col = "algorithm|name" if "algorithm|name" in all_cases_df.columns else None
        param_algo_col = "parameter|algorithm" if "parameter|algorithm" in all_cases_df.columns else None
        comparison_cols = [c for c in all_cases_df.columns
                           if "vs" in str(c) and "relative improvement" in str(c)]
        if not algo_name_col or not comparison_cols:
            logger.warning("Could not find required columns (algorithm|name or comparison columns)")
            return

        # Group rows into training/inference; split KNN by method (brute/kd_tree)
        fit_groups, inf_groups = {}, {}
        for algo_name in all_cases_df[algo_name_col].unique():
            adf = all_cases_df[all_cases_df[algo_name_col] == algo_name]
            target = (fit_groups if str(algo_name).endswith("|fit")
                      or str(algo_name) == "train_test_split" else inf_groups)
            if "kneighbors" in str(algo_name).lower() and param_algo_col:
                base, _, method = str(algo_name).partition("|")
                for pa in adf[param_algo_col].dropna().unique():
                    target[f"{base}({pa})|{method}"] = adf[adf[param_algo_col] == pa]
            else:
                target[algo_name] = adf

        color_fit, color_inference = '#004A99', '#E66100'

        def draw_panel(ax, groups, comp_col, title, color):
            labels, values = [], []
            for label, gdf in groups.items():
                vals = gdf[comp_col].dropna()
                if len(vals) > 0:
                    labels.append(label)
                    values.append(gmean(vals, nan_policy='omit'))
            if not values:
                return
            x = np.arange(len(labels))
            ax.set_axisbelow(True)
            bars = ax.bar(x, values, color=color, width=0.7, zorder=3)
            exp = min(4, max(1, int(np.floor(np.log10(max(values)))) + 1))
            y_ticks = [10 ** i for i in range(exp + 1)]
            ax.set_yscale('log')
            ax.yaxis.set_major_locator(FixedLocator(y_ticks))
            ax.yaxis.set_minor_locator(NullLocator())
            ax.yaxis.set_major_formatter(FixedFormatter([f'{float(t):.1f}' for t in y_ticks]))
            ax.set_ylim(1, y_ticks[-1])
            ax.grid(axis='y', which='major', linestyle='-', linewidth=0.8, color='#e0e0e0', zorder=0)
            ax.set_title(title, fontsize=16, color='#555555', pad=15)
            ax.set_ylabel('Speedup over original version\n(higher is better)', color='#555555', fontsize=11)
            ax.set_xlabel('scikit-learn* Algorithms', fontweight='bold', labelpad=10, fontsize=11)
            ax.set_xticks(x)
            ax.set_xticklabels([l.replace("|", " | ") for l in labels],
                               rotation=45, ha='right', fontsize=9)
            for spine in ('top', 'right'):
                ax.spines[spine].set_visible(False)
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width() / 2, height * 1.1, f'{height:.1f}',
                        ha='center', va='bottom', rotation=90, fontsize=8, color='#555555')

        # One row per comparison: training (left) + inference (right)
        n = len(comparison_cols)
        fig, axes = plt.subplots(n, 2, figsize=(16, 7 * n))
        if n == 1:
            axes = [axes]
        for i, comp_col in enumerate(comparison_cols):
            draw_panel(axes[i][0], fit_groups, comp_col, 'Training', color_fit)
            draw_panel(axes[i][1], inf_groups, comp_col, 'Inference', color_inference)

        # Reserve top margin for the two-component title
        plt.tight_layout(rect=[0, 0.06, 1, 0.90])

        # Two-component suptitle (black main title + blue subtitle)
        fig.text(
            0.5, 0.98,
            "Performance Benefits of Extension for Scikit-learn*",
            fontsize=22, color='#404040', ha='center', va='top',
        )
        fig.text(
            0.5, 0.935,
            "Combined Averages of FP32 & FP64 Workloads",
            fontsize=17, color='#0068B5', ha='center', va='top',
        )

        # Configuration placeholders — edit these to match the machine/run.
        TEST_DATE = "Month DD, YYYY"
        HARDWARE_CONFIG = (
            "x-node, CPU NAME, xx cores per socket, x sockets (x used), "
            "microcode xxx, HT on/off, Turbo on/off, SNC on/off (x NUMA nodes), "
            "xxxxGB (RAM type)"
        )
        SOFTWARE_CONFIG = (
            "Ubuntu xxx Python xxx "
            "Python libraries"
        )

        def mathbf(text):
            # Bold run for matplotlib mathtext (spaces must be escaped as '\ ')
            return r"$\bf{" + text.replace(" ", r"\ ") + "}$"

        # Footnote / disclaimer, drawn line by line so the URL renders as a link.
        # Each entry is (text, is_bold); "__URL__" is a special marker for the
        # line that embeds the www.Intel.com/PerformanceIndex link.
        footnote_lines = [
            (f"{mathbf('Testing Date:')} Performance results are based on "
             f"{mathbf(f'testing by Intel as of {TEST_DATE}')} and may not reflect all publically available security updates", False),
            (f"{mathbf('Configuration Details and Workload Setup:')} {HARDWARE_CONFIG}", False),
            (SOFTWARE_CONFIG, False),
            ('See backup for workloads and configurations. Performance results are based on testing as of dates shown in configurations ', False),
            ("__URL__", False),
            ("No product or component can be absolutely secure. Your costs and results may vary. Intel technologies may require enabled hardware, software or service activation.", False),
            ("© Intel Corporation. Intel, the Intel logo, and other Intel marks are trademarks of Intel Corporation or its subsidiaries. Other names and brands may be claimed as the property of others.", False),
        ]

        url_prefix = "and may not reflect all publicly available updates. Results may vary. Performance varies by use, configuration and other factors. Learn more at "
        url_text = "www.Intel.com/PerformanceIndex"
        url_suffix = "."

        # Render once so text extents can be measured for link placement
        fig.canvas.draw()
        renderer = fig.canvas.get_renderer()
        fig_w, fig_h = fig.bbox.width, fig.bbox.height

        x0, y0, step = 0.01, 0.05, 0.022
        gray, blue = "#606060", "#0068B5"
        y = y0
        for text, is_bold in footnote_lines:
            if text == "__URL__":
                t_pref = fig.text(x0, y, url_prefix, fontsize=9, color=gray,
                                  ha="left", va="top")
                w_pref = t_pref.get_window_extent(renderer=renderer).width / fig_w
                t_url = fig.text(x0 + w_pref, y, url_text, fontsize=9, color=blue,
                                 ha="left", va="top")
                ext = t_url.get_window_extent(renderer=renderer)
                x_start, x_end = ext.x0 / fig_w, ext.x1 / fig_w
                y_line = ext.y0 / fig_h
                fig.add_artist(plt.Line2D([x_start, x_end], [y_line, y_line],
                                          transform=fig.transFigure,
                                          color=blue, linewidth=0.8))
                fig.text(x_end, y, url_suffix, fontsize=9, color=gray,
                         ha="left", va="top")
            else:
                fig.text(x0, y, text, fontsize=9, color=gray, ha="left", va="top",
                         fontweight="bold" if is_bold else "normal")
            y -= step

        if output_file:
            plt.savefig(output_file, dpi=150, bbox_inches='tight')
            logger.info(f"Plot saved to {output_file}")
        else:
            plt.show()

        plt.close()

    except Exception as e:
        logger.error(f"Error drawing plots: {e}")


def generate_report(args: argparse.Namespace):
    logger.setLevel(args.report_log_level)
    results = merge_result_files(args.result_files)

    diffby, splitby = args.diff_columns, args.split_columns
    dfs = get_result_tables_as_df(
        results,
        diffby,
        splitby,
        args.compatibility_mode,
        args.performance_stability_metrics,
    )

    wb = xl.Workbook()
    summary_dfs = list()
    all_dfs = list()
    compared_dfs = dict()
    for df_name, df in dfs.items():
        drop_columns = list(set(df.columns) & set(args.drop_columns))
        df = df.drop(columns=drop_columns)

        ws = wb.create_sheet(title=df_name[:30])
        if len(diffby) > 0:
            current_df = compare_df(df, diffby, args.diffs_selection)
        else:
            current_df = df
        write_df_to_sheet(current_df, ws, index=False)
        apply_rules_for_sheet(ws, args.perf_color_scale, args.quality_color_scale)
        summary_dfs.append(get_summary_from_df(current_df, df_name))
        compared_dfs[df_name] = current_df
        # Add algorithm name column for tracking in all_cases sheet
        current_df_with_name = current_df.copy()
        current_df_with_name.insert(0, ("algorithm", "name"), df_name)
        all_dfs.append(current_df_with_name)
    # write summary to corresponding sheet
    all_cases_df = pd.concat(all_dfs, axis=0, join="outer")
    summary_df = pd.concat(summary_dfs, axis=0, join="outer")
    summary_df = summary_df[summary_df.columns.sortlevel(level=0, ascending=False)[0]]
    logger.info(f"{custom_format('Report summary', bcolor='HEADER')}\n{summary_df}")
    if summary_df.size > 0:
        summary_ws = wb.create_sheet(title="Summary", index=0)
        write_df_to_sheet(summary_df, summary_ws)
        apply_rules_for_sheet(summary_ws, args.perf_color_scale, args.quality_color_scale)
    if (all_cases_df.size > 0) and args.combined_results:
        # Prepare all_cases_df with proper column ordering (used for plots)
        all_cases_df = prepare_all_cases_df(all_cases_df)
        # Write "All cases" sheet with simplified format and GEOMEAN formulas
        geomean_rows = write_all_cases_2_sheet(compared_dfs, wb)
        # Write "Summary (for plots)" sheet referencing geomean values from "All cases"
        if geomean_rows:
            write_summary_2_sheet(geomean_rows, wb)
    # write environment info
    write_environment_info(results, wb)
    # remove default sheet
    wb.remove(wb["Sheet"])
    wb.save(args.report_file)
    
    # Draw plots if requested
    if args.draw_plots and (all_cases_df.size > 0):
        draw_summary_plots(all_cases_df, args.plot_output)
    
    return 0
